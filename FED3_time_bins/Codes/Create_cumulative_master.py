import os
import shutil
from numbers import Number
from openpyxl import load_workbook


# These are the only metrics that will be cumulatively summed.
# Percentages and averages are intentionally absent.
COMMON_SUM_METRICS = {
    "Number of blocks",
    "Length (mins)",
    "Left poke count",
    "Right poke count",
    "Total pokes",
    "Pellet count",
    "Sum of retrieval times (secs)",
    "Sum of IPIs (secs)",
    "Sum of poke times (secs)",
}


STOPSIG_SUM_METRICS = {
    ">Left_Regular_trial count",
    ">Left_Stop_trial count",
    "LeftinTimeOut count",
    "Right_no_left count",
    "RightDuringDispense count",
    "RightinTimeout count",

    "Regular LRP count",
    "Regular LRP latency LR sum (secs)",
    "Regular LRP latency RP sum (secs)",
    "Regular LN count",
    "Stop LNP count",
    "Stop LNP latency NP sum (secs)",
    "Stop LR count",
    "Stop LR latency LR sum (secs)",

    "Total regular events",
    "Total stop events",
    "Total events",
}


BANDIT_SUM_METRICS = {
    "Num reversals",
    "Pokes during dispense",
    "Poke in timeout",
    "Pokes with pellet",

    "Win",
    "High prob win",
    "Low prob win",

    "High prob win-stay",
    "High prob win-shift",
    "Low prob win-stay",
    "Low prob win-shift",

    "Loss",
    "High prob loss",
    "Low prob loss",

    "High prob lose-stay",
    "High prob lose-shift",
    "Low prob lose-stay",
    "Low prob lose-shift",
}


SESSION_SUM_METRICS = {
    "StopSig": STOPSIG_SUM_METRICS,
    "LeftRight": STOPSIG_SUM_METRICS,
    "ClosedEcon_PR1": set(),
    "Bandit": BANDIT_SUM_METRICS,
}


def normalise_name(value):
    """
    Normalise a metric name for case-insensitive matching.
    """
    if value is None:
        return ""

    return " ".join(str(value).strip().casefold().split())


def get_sum_metrics(session_type):
    """
    Return the approved additive metrics for a session type.
    """
    metrics = (
        COMMON_SUM_METRICS
        | SESSION_SUM_METRICS.get(session_type, set())
    )

    return {normalise_name(metric) for metric in metrics}


def is_number(value):
    """
    True for Excel numeric cells, excluding Boolean values.
    """
    return isinstance(value, Number) and not isinstance(value, bool)


def add_sum_suffix(file_path):
    """
    Convert Example.xlsx to Example_Sum.xlsx.
    """
    base, extension = os.path.splitext(file_path)
    return f"{base}_Sum{extension}"


def has_later_numeric_value(
    ws,
    start_row,
    column_number,
    filename_column=None,
    filename=None,
):
    """
    Check whether the same column has another numeric value later.

    When filename_column is supplied, only inspect rows belonging
    to the same filename.
    """
    for row_number in range(start_row + 1, ws.max_row + 1):
        if filename_column is not None:
            later_filename = ws.cell(
                row=row_number,
                column=filename_column
            ).value

            if later_filename != filename:
                continue

        later_value = ws.cell(
            row=row_number,
            column=column_number
        ).value

        if is_number(later_value):
            return True

    return False


def find_first_data_row(ws):
    """
    Find the first numeric entry in column A.

    In the metric-sheet workbooks, column A contains the block,
    cycle, or day index. Therefore its first numeric value marks
    the beginning of the actual data.
    """
    for row_number in range(1, ws.max_row + 1):
        value = ws.cell(row=row_number, column=1).value

        if is_number(value):
            return row_number

    return None


def process_metric_sheet(ws):
    """
    Apply a running sum down every data column in one metric sheet.

    Column A is an identifying block/day/cycle column, so it is
    deliberately left unchanged.
    """
    first_data_row = find_first_data_row(ws)

    if first_data_row is None:
        return

    for column_number in range(2, ws.max_column + 1):
        running_sum = 0
        found_number = False

        for row_number in range(first_data_row, ws.max_row + 1):
            cell = ws.cell(
                row=row_number,
                column=column_number
            )
            value = cell.value

            if is_number(value):
                running_sum += value
                cell.value = running_sum
                found_number = True

            elif (
                value is None
                and found_number
                and has_later_numeric_value(
                    ws,
                    row_number,
                    column_number,
                )
            ):
                # Match Longfix behaviour for internal blank cells.
                cell.value = running_sum


def find_header_columns(ws):
    """
    Read the first row of a combined master sheet and return a
    normalised heading-to-column-number dictionary.
    """
    columns = {}

    for column_number in range(1, ws.max_column + 1):
        heading = ws.cell(
            row=1,
            column=column_number
        ).value

        if heading is not None:
            columns[normalise_name(heading)] = column_number

    return columns


def process_combined_sheet(ws, sum_metrics):
    """
    Process a combined master sheet.

    These sheets contain several metric columns and several rows
    for each filename. Each cumulative sum restarts separately for
    each filename.
    """
    header_columns = find_header_columns(ws)
    filename_column = header_columns.get("filename")

    if filename_column is None:
        return

    metric_columns = {
        heading: column_number
        for heading, column_number in header_columns.items()
        if heading in sum_metrics
    }

    for column_number in metric_columns.values():
        running_totals = {}
        found_numbers = set()

        for row_number in range(2, ws.max_row + 1):
            filename = ws.cell(
                row=row_number,
                column=filename_column
            ).value

            if filename is None:
                continue

            cell = ws.cell(
                row=row_number,
                column=column_number
            )
            value = cell.value

            if is_number(value):
                running_totals[filename] = (
                    running_totals.get(filename, 0) + value
                )
                cell.value = running_totals[filename]
                found_numbers.add(filename)

            elif (
                value is None
                and filename in found_numbers
                and has_later_numeric_value(
                    ws,
                    row_number,
                    column_number,
                    filename_column=filename_column,
                    filename=filename,
                )
            ):
                # Match Longfix behaviour for internal blank cells.
                cell.value = running_totals[filename]


def create_sum_workbook(
    source_path,
    session_type,
    layout,
):
    """
    Copy an existing master workbook and create its _Sum version.

    layout must be:
        "combined"
        "metric_sheets"
    """
    if not os.path.exists(source_path):
        raise FileNotFoundError(
            f"Master workbook was not found: {source_path}"
        )

    output_path = add_sum_suffix(source_path)

    # Copy first so the original formatting is preserved.
    shutil.copy2(source_path, output_path)

    workbook = load_workbook(output_path)
    sum_metrics = get_sum_metrics(session_type)

    if layout == "combined":
        for worksheet in workbook.worksheets:
            process_combined_sheet(
                worksheet,
                sum_metrics,
            )

    elif layout == "metric_sheets":
        kept_sheet_count = 0

        # Use list(...) because worksheets may be removed while looping.
        for worksheet in list(workbook.worksheets):
            sheet_name = normalise_name(worksheet.title)

            if sheet_name in sum_metrics:
                process_metric_sheet(worksheet)
                kept_sheet_count += 1
            else:
                workbook.remove(worksheet)

        # Excel workbooks must contain at least one worksheet.
        # If no approved additive metrics were found, do not create an empty _Sum workbook.
        if kept_sheet_count == 0:
            workbook.close()

            if os.path.exists(output_path):
                os.remove(output_path)

            return None

    else:
        workbook.close()
        raise ValueError(
            f"Unknown cumulative workbook layout: {layout}"
        )

    workbook.save(output_path)
    workbook.close()

    return output_path